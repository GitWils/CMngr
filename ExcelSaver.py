import openpyxl
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill

from pprint import pprint

class ExcelSaver():
    def __init__(self, filename):
        print(f"saved in: {filename}")
        self.wb = openpyxl.Workbook()
        self.filename = filename
    def writeComponents(self, components, offset = 0):
        """ write components excel sheet """
        sheet = self.wb.active
        self.setStyleTableNameCell(sheet.cell(1, 1 + offset), 'Надходження комплектуючих')
        self.setStyleHeaderCell(sheet.cell(3, 1 + offset), 'Виріб')
        self.setStyleHeaderCell(sheet.cell(3, 2 + offset), 'Деталь')
        self.setStyleHeaderCell(sheet.cell(3, 3 + offset), 'Договір')
        self.setStyleHeaderCell(sheet.cell(3, 4 + offset), 'Кількість (шт. або кг)')
        self.setStyleHeaderCell(sheet.cell(3, 5 + offset), 'Дата надходження')
        self.setStyleHeaderCell(sheet.cell(3, 6 + offset), 'Примітка')

        for i in range (0, len(components)):
            self.setStyleCell(sheet.cell(i + 4, 1 + offset), components[i]['name'])
            self.setStyleCell(sheet.cell(i + 4, 2 + offset), components[i]['device'])
            self.setStyleCell(sheet.cell(i + 4, 3 + offset), components[i]['contract'])
            self.setStyleCell(sheet.cell(i + 4, 4 + offset), components[i]['count'])
            self.setStyleCell(sheet.cell(i + 4, 5 + offset), components[i]['date'][5:])
            self.setStyleCell(sheet.cell(i + 4, 6 + offset), components[i]['note'])
        self.alignSheet(sheet)

    def writeReports(self, reports, offset = 0):
        """ write reports to excel sheet """
        sheet = self.wb.active
        self.setStyleTableNameCell(sheet.cell(1, 1 + offset), 'Звітні дані')
        self.setStyleHeaderCell(sheet.cell(3, 1 + offset), 'Виріб')
        self.setStyleHeaderCell(sheet.cell(3, 2 + offset), 'Деталь')
        self.setStyleHeaderCell(sheet.cell(3, 3 + offset), 'Договір')
        self.setStyleHeaderCell(sheet.cell(3, 4 + offset), 'Наявність (шт. або кг)')
        self.setStyleHeaderCell(sheet.cell(3, 5 + offset), 'Очікується')
        self.setStyleHeaderCell(sheet.cell(3, 6 + offset), 'Всього необхідно')

        for i in range (0, len(reports)):
            self.setStyleCell(sheet.cell(i + 4, 1 + offset), reports[i]['product'])
            self.setStyleCell(sheet.cell(i + 4, 2 + offset), reports[i]['device'])
            self.setStyleCell(sheet.cell(i + 4, 3 + offset), reports[i]['contract'])
            self.setStyleCell(sheet.cell(i + 4, 4 + offset), reports[i]['not_assembled'])
            self.setStyleCell(sheet.cell(i + 4, 5 + offset), reports[i]['needed'])
            self.setStyleCell(sheet.cell(i + 4, 6 + offset), reports[i]['not_assembled'] - reports[i]['needed'])
        self.alignSheet(sheet)

    def writeShipments(self, shipments, offset = 0):
        """ write shipments excel sheet """
        sheet = self.wb.active
        self.setStyleTableNameCell(sheet.cell(1, 1 + offset), 'Відвантажено')
        self.setStyleHeaderCell(sheet.cell(3, 1 + offset), 'Виріб')
        self.setStyleHeaderCell(sheet.cell(3, 2 + offset), 'Договір')
        self.setStyleHeaderCell(sheet.cell(3, 3 + offset), 'Відвантажено')
        self.setStyleHeaderCell(sheet.cell(3, 4 + offset), 'Всього по договору')
        self.setStyleHeaderCell(sheet.cell(3, 5 + offset), 'Дата відвантаження')
        self.setStyleHeaderCell(sheet.cell(3, 6 + offset), 'Примітка')

        for i in range (0, len(shipments)):
            self.setStyleCell(sheet.cell(i + 4, 1 + offset), shipments[i]['product'])
            self.setStyleCell(sheet.cell(i + 4, 2 + offset), shipments[i]['contract_name'])
            self.setStyleCell(sheet.cell(i + 4, 3 + offset), shipments[i]['sended'])
            self.setStyleCell(sheet.cell(i + 4, 4 + offset), shipments[i]['count'])
            self.setStyleCell(sheet.cell(i + 4, 5 + offset), shipments[i]['str_date'][5:])
            self.setStyleCell(sheet.cell(i + 4, 6 + offset), shipments[i]['note'])
        self.alignSheet(sheet)

    def writeAssemblings(self, assemblings, offset = 0):
        """ write shipments excel sheet """
        sheet = self.wb.active
        self.setStyleTableNameCell(sheet.cell(1, 1 + offset), 'Зібрано')
        self.setStyleHeaderCell(sheet.cell(3, 1 + offset), 'Виріб')
        self.setStyleHeaderCell(sheet.cell(3, 2 + offset), 'Договір')
        self.setStyleHeaderCell(sheet.cell(3, 3 + offset), 'Зібрано')
        self.setStyleHeaderCell(sheet.cell(3, 4 + offset), 'Примітка')
        self.setStyleHeaderCell(sheet.cell(3, 5 + offset), 'Дата зборки')

        for i in range(0, len(assemblings)):
            self.setStyleCell(sheet.cell(i + 4, 1 + offset), assemblings[i]['name'])
            self.setStyleCell(sheet.cell(i + 4, 2 + offset), assemblings[i]['contract_name'])
            self.setStyleCell(sheet.cell(i + 4, 3 + offset), assemblings[i]['count'])
            self.setStyleCell(sheet.cell(i + 4, 4 + offset), assemblings[i]['note'])
            self.setStyleCell(sheet.cell(i + 4, 5 + offset), assemblings[i]['date'][5:])

    def saveToFile(self):
        """ write data to excel file """
        if self.filename.endswith(".xlsx"):
            self.wb.save(self.filename)
        else:
            self.wb.save(self.filename + ".xlsx")

    def setStyleTableNameCell(self, cell, value):
        """ initialisation table name styles """
        cell.value = value
        cell.alignment = Alignment(horizontal='left', vertical='center')
        cell.font = Font(bold = 'True', size = 16)

    def setStyleHeaderCell(self, cell, value, bold = 'True'):
        """ initialisation header table styles """
        cell.value = value
        cell.font = Font(bold = bold, size = 11, color = '003300')
        cell.fill = PatternFill("solid", fgColor="dde8cb")
        cell.alignment = Alignment(horizontal='center', vertical='center')
        style = Side(border_style = 'thin', color = '555555')
        cell.border = Border(bottom = style, top = style, left = style, right = style)

    def setStyleCell(self, cell, value):
        """ initialisation table cell styles """
        cell.value = value
        cell.alignment = Alignment(horizontal='center', vertical='center')

    def alignSheet(self, sheet):
        """ resize columns to content """
        for column_cells in sheet.iter_cols(min_col=1, max_col=sheet.max_column):
            max_length = 0
            column = column_cells[0].column  # Отримуємо номер стовпця
            for cell in column_cells:
                if cell.font.sz < 16 and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            #do increase cell width by 10% of max_length
            sheet.column_dimensions[openpyxl.utils.get_column_letter(column)].width = round((max_length + 2) * 1.1)